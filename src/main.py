import time
import os
import sys
import platform
from pathlib import Path
from openpyxl import Workbook, load_workbook
from core.exctract_data import extract_pdfs
from core.proveniencia import Proveniencia

sys.stdout.reconfigure(encoding='utf-8')

# Raiz do projeto = pasta acima de src/ — funciona em qualquer máquina.
PROJETO_RAIZ = Path(__file__).resolve().parent.parent
DATA_DIR = PROJETO_RAIZ / "Data"

leitura_directory = str(DATA_DIR / "input")
Excel_Output = DATA_DIR / "output" / "Teste.xlsx"
caminho_db = DATA_DIR / "proveniencia.db"

inicio = time.time()

files = os.listdir(leitura_directory)

if not files:
    raise Exception("Nenhum arquivo encontrado nesse diretório")

if os.path.exists(Excel_Output):
    wb = load_workbook(Excel_Output)
    ws = wb.active

    if "Erros" in wb.sheetnames:
        ws_erro = wb["Erros"]
    else:
        ws_erro = wb.create_sheet("Erros")
        ws_erro.append(["Banco", "Arquivo", "Página", "Linha", "Conteúdo", "Erro"])
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Combinado"
    ws.append([
        "Data", "Conta Corrente", "Histórico", "Dcto", "Valor", "D/C Valor",
        "Saldo", "D/C Saldo", "Banco", "Nome do PDF", "Pagina", "Linha", "Status"
    ])
    ws_erro = wb.create_sheet("Erros")
    ws_erro.append(["Banco", "Arquivo", "Página", "Linha", "Conteúdo", "Erro"])

prov = Proveniencia(caminho_db)
execucao_id = None
resultado = {"total_sucesso": 0, "total_com_erro": 0, "total_registros": 0}
status = "FALHA"
mensagem_erro = None

try:
    try:
        execucao_id = prov.iniciar_execucao(leitura_directory, str(Excel_Output), len(files))
    except Exception as e:
        print(f"[proveniencia] falha ao iniciar execução: {e}")

    resultado = extract_pdfs(files, leitura_directory, ws, ws_erro, prov, execucao_id)

    if resultado["total_com_erro"] == 0:
        status = "SUCESSO"
    elif resultado["total_sucesso"] > 0:
        status = "SUCESSO_PARCIAL"
    else:
        status = "FALHA"
except Exception as e:
    mensagem_erro = str(e)
    status = "FALHA"
    raise
finally:
    wb.save(Excel_Output)

    fim = time.time()
    duracao = fim - inicio

    if execucao_id is not None:
        try:
            prov.finalizar_execucao(
                execucao_id, duracao,
                resultado["total_sucesso"], resultado["total_com_erro"],
                resultado["total_registros"], status, mensagem_erro
            )
        except Exception as e:
            print(f"[proveniencia] falha ao finalizar execução: {e}")

    prov.fechar()
    print(f"Tempo total: {duracao/60:.2f} minutos")
